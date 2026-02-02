import os
import asyncio
import re
import logging
import sys
from datetime import datetime, timedelta
from telethon import TelegramClient, events
from telethon.sessions import StringSession
from aiohttp import web
import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import (
    API_ID, API_HASH, BOT_TOKEN, ADMIN_ID,
    SOURCE_CHANNEL_ID, BILAN_CHANNEL_ID, PORT,
    BILAN_INTERVAL_MINUTES, DATABASE_URL,
    JOUR_START, JOUR_END
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

client = TelegramClient(StringSession(''), API_ID, API_HASH)
bilan_interval = BILAN_INTERVAL_MINUTES
current_jour_id = None  # Identifiant de la journée actuelle (format: YYYY-MM-DD)

class PostgresDB:
    def __init__(self, database_url):
        self.database_url = database_url
        self.pool = None

    async def connect(self):
        try:
            self.pool = await asyncpg.create_pool(self.database_url, min_size=1, max_size=10)
            logger.info("Connecté à PostgreSQL")
            await self.create_tables()
        except Exception as e:
            logger.error(f"Erreur connexion PostgreSQL: {e}")
            raise

    async def create_tables(self):
        async with self.pool.acquire() as conn:
            # Table des jeux avec jour_id
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS games (
                    id SERIAL PRIMARY KEY,
                    jour_id VARCHAR(20) NOT NULL,
                    game_number INTEGER NOT NULL,
                    suit VARCHAR(10) NOT NULL,
                    category VARCHAR(10) NOT NULL,
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    raw_line TEXT,
                    UNIQUE(jour_id, game_number)
                )
            """)

            # Table des jours (6-1436)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS jours (
                    id SERIAL PRIMARY KEY,
                    jour_id VARCHAR(20) UNIQUE NOT NULL,
                    date_str VARCHAR(20) NOT NULL,
                    start_num INTEGER DEFAULT 6,
                    end_num INTEGER DEFAULT 1436,
                    is_complete BOOLEAN DEFAULT FALSE,
                    total_games INTEGER DEFAULT 0,
                    count_0 INTEGER DEFAULT 0,
                    count_1 INTEGER DEFAULT 0,
                    count_2 INTEGER DEFAULT 0,
                    count_3 INTEGER DEFAULT 0,
                    count_loss INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Table des numéros (indépendante des jours)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS number_stats (
                    id SERIAL PRIMARY KEY,
                    number INTEGER UNIQUE NOT NULL,
                    appearances INTEGER DEFAULT 0,
                    count_0 INTEGER DEFAULT 0,
                    count_1 INTEGER DEFAULT 0,
                    count_2 INTEGER DEFAULT 0,
                    count_3 INTEGER DEFAULT 0,
                    count_loss INTEGER DEFAULT 0,
                    first_seen TIMESTAMP,
                    last_seen TIMESTAMP,
                    has_never_lost BOOLEAN DEFAULT TRUE
                )
            """)

            # Table pour les comparaisons entre jours
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS day_comparisons (
                    id SERIAL PRIMARY KEY,
                    jour_id_1 VARCHAR(20) NOT NULL,
                    jour_id_2 VARCHAR(20) NOT NULL,
                    common_numbers INTEGER[],
                    common_cat_0 INTEGER[],
                    common_cat_1 INTEGER[],
                    common_cat_2 INTEGER[],
                    common_cat_3 INTEGER[],
                    common_loss INTEGER[],
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(jour_id_1, jour_id_2)
                )
            """)

            logger.info("Tables PostgreSQL créées")

    async def get_or_create_jour(self, game_number):
        """Détermine ou crée le jour actuel basé sur le numéro de jeu"""
        global current_jour_id

        today = datetime.now()

        # Si on est entre 6 et 1436, c'est la journée normale
        if JOUR_START <= game_number <= JOUR_END:
            jour_id = today.strftime("%Y-%m-%d")
        else:
            # Si c'est 1-5 ou 1437-1440, c'est la journée précédente ou suivante
            # Pour simplifier, on utilise la date du jour mais on note l'exception
            if game_number < JOUR_START:
                # Fait partie de la journée précédente
                yesterday = today - timedelta(days=1)
                jour_id = yesterday.strftime("%Y-%m-%d")
            else:
                # 1437-1440 fait partie du jour suivant (rare)
                tomorrow = today + timedelta(days=1)
                jour_id = tomorrow.strftime("%Y-%m-%d")

        current_jour_id = jour_id

        async with self.pool.acquire() as conn:
            # Vérifier si le jour existe
            exists = await conn.fetchval("""
                SELECT 1 FROM jours WHERE jour_id = $1
            """, jour_id)

            if not exists:
                await conn.execute("""
                    INSERT INTO jours (jour_id, date_str, start_num, end_num)
                    VALUES ($1, $2, $3, $4)
                """, jour_id, jour_id, JOUR_START, JOUR_END)
                logger.info(f"Nouvelle journée créée: {jour_id} (jeux {JOUR_START}-{JOUR_END})")

        return jour_id

    async def add_game(self, game_number, suit, category, raw_line):
        jour_id = await self.get_or_create_jour(game_number)

        async with self.pool.acquire() as conn:
            # Insérer le jeu
            await conn.execute("""
                INSERT INTO games (jour_id, game_number, suit, category, raw_line)
                VALUES ($1, $2, $3, $4, $5)
                ON CONFLICT (jour_id, game_number) DO UPDATE SET
                    suit = EXCLUDED.suit,
                    category = EXCLUDED.category,
                    raw_line = EXCLUDED.raw_line
            """, jour_id, game_number, suit, category, raw_line)

            # Mettre à jour les stats du jour
            cat_col = {
                '✅0️⃣': 'count_0',
                '✅1️⃣': 'count_1',
                '✅2️⃣': 'count_2',
                '✅3️⃣': 'count_3',
                '❌': 'count_loss'
            }.get(category, None)

            if cat_col:
                await conn.execute(f"""
                    UPDATE jours 
                    SET total_games = total_games + 1,
                        {cat_col} = {cat_col} + 1
                    WHERE jour_id = $1
                """, jour_id)

            # Vérifier si journée complète (1436 atteint)
            if game_number == JOUR_END:
                await conn.execute("""
                    UPDATE jours SET is_complete = TRUE WHERE jour_id = $1
                """, jour_id)
                logger.info(f"Journée {jour_id} marquée comme complète")

            # Mettre à jour les stats globales du numéro
            await conn.execute(f"""
                INSERT INTO number_stats (number, appearances, {cat_col}, first_seen, last_seen, has_never_lost)
                VALUES ($1, 1, 1, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, $2)
                ON CONFLICT (number) DO UPDATE SET
                    appearances = number_stats.appearances + 1,
                    {cat_col} = number_stats.{cat_col} + 1,
                    last_seen = CURRENT_TIMESTAMP,
                    has_never_lost = CASE WHEN $3 = '❌' THEN FALSE ELSE number_stats.has_never_lost END
            """, game_number, category != '❌', category)

    async def get_jour_stats(self, jour_id=None):
        if jour_id is None:
            jour_id = current_jour_id

        async with self.pool.acquire() as conn:
            row = await conn.fetchrow("""
                SELECT * FROM jours WHERE jour_id = $1
            """, jour_id)
            return row

    async def get_numbers_by_category_and_jour(self, category, jour_id=None):
        if jour_id is None:
            jour_id = current_jour_id

        async with self.pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT game_number, suit, timestamp FROM games 
                WHERE jour_id = $1 AND category = $2
                ORDER BY game_number
            """, jour_id, category)
            return rows

    async def get_all_games_by_jour(self, jour_id=None):
        if jour_id is None:
            jour_id = current_jour_id

        async with self.pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT * FROM games WHERE jour_id = $1 ORDER BY game_number
            """, jour_id)
            return rows

    async def get_all_jours(self):
        async with self.pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT * FROM jours ORDER BY jour_id DESC
            """)
            return rows

    async def get_comparison_data(self, jour_id_1, jour_id_2):
        """Compare deux journées et retourne les numéros communs par catégorie"""
        async with self.pool.acquire() as conn:
            # Numéros communs
            common = await conn.fetch("""
                SELECT g1.game_number, g1.category
                FROM games g1
                INNER JOIN games g2 ON g1.game_number = g2.game_number AND g1.category = g2.category
                WHERE g1.jour_id = $1 AND g2.jour_id = $2
                ORDER BY g1.game_number
            """, jour_id_1, jour_id_2)

            return common

db = PostgresDB(DATABASE_URL)

# Fonctions d'Analyse
def parse_game_message(message_text):
    games = []
    lines = message_text.strip().split('
')

    for line in lines:
        line = line.strip()
        if not line or '—' not in line:
            continue

        number_match = re.match(r'(\d+)\s*—', line)
        if not number_match:
            continue

        game_number = int(number_match.group(1))

        suit_match = re.search(r'игрок\s*([♠♥♦♣❤️♠️♥️♦️♣️])', line)
        if not suit_match:
            continue

        suit = suit_match.group(1)
        suit = suit.replace('❤️', '♥️').replace('❤', '♥️').replace('♥', '♥️')
        suit = suit.replace('♠', '♠️').replace('♦', '♦️').replace('♣', '♣️')

        category = None
        if '✅ 0️⃣' in line or '✅0️⃣' in line:
            category = '✅0️⃣'
        elif '✅ 1️⃣' in line or '✅1️⃣' in line:
            category = '✅1️⃣'
        elif '✅ 2️⃣' in line or '✅2️⃣' in line:
            category = '✅2️⃣'
        elif '✅ 3️⃣' in line or '✅3️⃣' in line:
            category = '✅3️⃣'
        elif '❌' in line:
            category = '❌'

        if category:
            games.append({
                'number': game_number,
                'suit': suit,
                'category': category,
                'raw_line': line
            })

    return games

# Envoi Automatique des Bilans
async def send_bilan():
    try:
        stats = await db.get_jour_stats()

        if not stats:
            logger.info("Aucune donnée pour le bilan")
            return

        today_str = datetime.now().strftime("%d/%m/%Y")

        msg = f"""📊 **BILAN AUTOMATIQUE - {today_str}**

🎮 **Journée:** {stats['jour_id']}
📊 **Jeux {JOUR_START}-{JOUR_END}:** {stats['total_games']}
✅ **Complète:** {"Oui" if stats['is_complete'] else "En cours"}

**Répartition:**
• ✅0️⃣: {stats['count_0']} jeux
• ✅1️⃣: {stats['count_1']} jeux
• ✅2️⃣: {stats['count_2']} jeux  
• ✅3️⃣: {stats['count_3']} jeux
• ❌: {stats['count_loss']} jeux

⏰ Prochain bilan dans {bilan_interval} minutes
"""

        await client.send_message(BILAN_CHANNEL_ID, msg)
        logger.info(f"Bilan envoyé au canal {BILAN_CHANNEL_ID}")

    except Exception as e:
        logger.error(f"Erreur envoi bilan: {e}")

async def bilan_scheduler():
    while True:
        await asyncio.sleep(bilan_interval * 60)
        await send_bilan()

# Gestion des Messages
async def process_edited_message(message_text, chat_id):
    try:
        if chat_id != SOURCE_CHANNEL_ID:
            return

        games = parse_game_message(message_text)

        if not games:
            return

        logger.info(f"{len(games)} jeux détectés dans le message édité")

        for game in games:
            await db.add_game(
                game_number=game['number'],
                suit=game['suit'],
                category=game['category'],
                raw_line=game['raw_line']
            )
            logger.info(f"Jeu #{game['number']} enregistré: {game['suit']} - {game['category']}")

    except Exception as e:
        logger.error(f"Erreur traitement message: {e}")

async def handle_edited_message(event):
    try:
        chat = await event.get_chat()
        chat_id = chat.id

        if hasattr(chat, 'broadcast') and chat.broadcast:
            if not str(chat_id).startswith('-100'):
                chat_id = int(f"-100{abs(chat_id)}")

        if chat_id == SOURCE_CHANNEL_ID:
            message_text = event.message.message
            logger.info(f"Message édité détecté du canal {chat_id}")
            await process_edited_message(message_text, chat_id)

    except Exception as e:
        logger.error(f"Erreur handle_edited_message: {e}")

# Commandes
@client.on(events.NewMessage(pattern='/start'))
async def cmd_start(event):
    if event.is_group or event.is_channel:
        return

    help_text = f"""🤖 **Bot de Collecte Baccarat**

**Configuration:**
• Journée: {JOUR_START} à {JOUR_END}
• Bilan: Toutes les {bilan_interval} min

**Commandes:**
• `/info` - Bilan du jour en cours
• `/set_interval <min>` - Changer intervalle
• `/force_bilan` - Envoyer bilan maintenant
• `/inter` - Export Excel complet
• `/inter_jour <JJ-MM-AAAA>` - Export d'un jour spécifique
"""
    await event.respond(help_text)

@client.on(events.NewMessage(pattern=r'/set_interval\s+(\d+)'))
async def cmd_set_interval(event):
    if event.is_group or event.is_channel:
        return

    global bilan_interval
    try:
        new_interval = int(event.pattern_match.group(1))
        if new_interval < 1:
            await event.respond("❌ Minimum 1 minute")
            return

        bilan_interval = new_interval
        await event.respond(f"✅ Intervalle: **{bilan_interval} minutes**")

    except Exception as e:
        await event.respond(f"❌ Erreur: {e}")

@client.on(events.NewMessage(pattern='/force_bilan'))
async def cmd_force_bilan(event):
    if event.is_group or event.is_channel:
        return

    await event.respond("📊 Envoi du bilan...")
    await send_bilan()
    await event.respond("✅ Bilan envoyé!")

@client.on(events.NewMessage(pattern='/info'))
async def cmd_info(event):
    if event.is_group or event.is_channel:
        return

    stats = await db.get_jour_stats()

    if not stats:
        await event.respond("❌ Aucune donnée")
        return

    msg = f"""📊 **Journée {stats['jour_id']}**

🎮 **Jeux {JOUR_START}-{JOUR_END}:** {stats['total_games']}
⏱ **Bilan:** {bilan_interval} min
✅ **Complète:** {"Oui" if stats['is_complete'] else "Non"}

**Répartition:**
• ✅0️⃣: {stats['count_0']}
• ✅1️⃣: {stats['count_1']}
• ✅2️⃣: {stats['count_2']}
• ✅3️⃣: {stats['count_3']}
• ❌: {stats['count_loss']}
"""
    await event.respond(msg)

# Export Excel avec feuilles par catégorie
async def create_excel_export(jour_id=None, filename=None):
    """Crée un fichier Excel avec une feuille par catégorie"""

    if jour_id is None:
        jour_id = current_jour_id

    if filename is None:
        filename = f"baccarat_{jour_id}.xlsx"

    # Récupérer les données
    categories = ['✅0️⃣', '✅1️⃣', '✅2️⃣', '✅3️⃣', '❌']
    cat_names = {'✅0️⃣': 'CAT_0', '✅1️⃣': 'CAT_1', '✅2️⃣': 'CAT_2', '✅3️⃣': 'CAT_3', '❌': 'CAT_LOSS'}

    wb = Workbook()

    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # Supprimer la feuille par défaut
    wb.remove(wb.active)

    # Créer une feuille par catégorie
    for cat in categories:
        ws = wb.create_sheet(title=cat_names[cat])

        # En-têtes
        ws['A1'] = "NUMÉRO"
        ws['B1'] = "COSTUME"
        ws['C1'] = "HEURE"

        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}1']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Récupérer les données de cette catégorie pour ce jour
        rows = await db.get_numbers_by_category_and_jour(cat, jour_id)

        # Remplir les données
        for idx, row in enumerate(rows, 2):
            ws[f'A{idx}'] = row['game_number']
            ws[f'B{idx}'] = row['suit']
            ws[f'C{idx}'] = row['timestamp'].strftime("%H:%M:%S")

        # Ajuster largeurs
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12

    # Feuille récapitulative
    ws_recap = wb.create_sheet(title="RÉCAP", index=0)
    ws_recap['A1'] = "CATÉGORIE"
    ws_recap['B1'] = "TOTAL"
    ws_recap['C1'] = "POURCENTAGE"

    for col in ['A', 'B', 'C']:
        cell = ws_recap[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Stats du jour
    stats = await db.get_jour_stats(jour_id)
    if stats:
        total = stats['total_games'] or 1
        data_recap = [
            ('✅0️⃣', stats['count_0']),
            ('✅1️⃣', stats['count_1']),
            ('✅2️⃣', stats['count_2']),
            ('✅3️⃣', stats['count_3']),
            ('❌', stats['count_loss'])
        ]

        for idx, (cat, count) in enumerate(data_recap, 2):
            ws_recap[f'A{idx}'] = cat
            ws_recap[f'B{idx}'] = count
            ws_recap[f'C{idx}'] = f"{(count/total)*100:.1f}%"

    ws_recap.column_dimensions['A'].width = 15
    ws_recap.column_dimensions['B'].width = 12
    ws_recap.column_dimensions['C'].width = 15

    wb.save(filename)
    return filename

@client.on(events.NewMessage(pattern='/inter'))
async def cmd_inter(event):
    if event.is_group or event.is_channel:
        return

    await event.respond("📁 Création de l'export Excel...")

    try:
        filename = await create_excel_export()
        await client.send_file(event.chat_id, filename, caption=f"📊 Export du jour ({current_jour_id})")
        os.remove(filename)

    except Exception as e:
        logger.error(f"Erreur export: {e}")
        await event.respond(f"❌ Erreur: {e}")

@client.on(events.NewMessage(pattern=r'/inter_jour\s+(\d{2}-\d{2}-\d{4})'))
async def cmd_inter_jour(event):
    if event.is_group or event.is_channel:
        return

    try:
        date_str = event.pattern_match.group(1)
        day, month, year = date_str.split('-')
        jour_id = f"{year}-{month}-{day}"

        await event.respond(f"📁 Export du {jour_id}...")

        filename = f"baccarat_{jour_id}.xlsx"
        await create_excel_export(jour_id, filename)

        await client.send_file(event.chat_id, filename, caption=f"📊 Export du {jour_id}")
        os.remove(filename)

    except Exception as e:
        await event.respond(f"❌ Erreur: {e}")


# Fonctions de comparaison globale
async def get_global_comparison_data():
    """Récupère les données pour comparer toutes les journées"""
    async with db.pool.acquire() as conn:
        # Tous les jours enregistrés
        jours = await conn.fetch("""
            SELECT jour_id FROM jours ORDER BY jour_id
        """)

        # Pour chaque numéro, dans combien de jours il apparaît et avec quelles catégories
        number_freq = await conn.fetch("""
            SELECT 
                game_number,
                COUNT(DISTINCT jour_id) as nb_jours,
                array_agg(DISTINCT jour_id) as jours,
                array_agg(DISTINCT category) as categories
            FROM games
            GROUP BY game_number
            ORDER BY nb_jours DESC, game_number
        """)

        # Numéros qui apparaissent dans tous les jours (avec même catégorie)
        common_numbers = await conn.fetch("""
            SELECT game_number, category, COUNT(DISTINCT jour_id) as freq
            FROM games
            GROUP BY game_number, category
            HAVING COUNT(DISTINCT jour_id) = (SELECT COUNT(*) FROM jours)
            ORDER BY game_number
        """)

        return {
            'jours': [j['jour_id'] for j in jours],
            'number_freq': number_freq,
            'common_numbers': common_numbers
        }


async def create_comparison_only_excel():
    """Crée un Excel UNIQUEMENT avec les comparaisons (pas de données brutes)"""

    async with db.pool.acquire() as conn:
        # Tous les jours
        jours = await conn.fetch("SELECT jour_id FROM jours ORDER BY jour_id")
        jours_list = [j['jour_id'] for j in jours]

        if len(jours_list) < 2:
            return None, "Minimum 2 journées requises pour comparer"

        wb = Workbook()
        wb.remove(wb.active)

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=13, color="1F4E78")

        # === FEUILLE 1: RÉCAP GLOBAL ===
        ws_recap = wb.create_sheet("RÉCAP GLOBAL", 0)

        ws_recap['A1'] = "ANALYSE COMPARATIVE - TOUTES LES JOURNÉES"
        ws_recap['A1'].font = Font(bold=True, size=14, color="1F4E78")
        ws_recap.merge_cells('A1:E1')

        ws_recap['A3'] = "Nombre total de journées:"
        ws_recap['B3'] = len(jours_list)
        ws_recap['B3'].font = Font(bold=True, size=12)

        ws_recap['A5'] = "Liste des journées analysées:"
        for idx, jour in enumerate(jours_list, 6):
            ws_recap[f'A{idx}'] = jour

        # Largeurs
        ws_recap.column_dimensions['A'].width = 35
        ws_recap.column_dimensions['B'].width = 15

        # === FEUILLE 2: NUMÉROS PAR FRÉQUENCE ===
        ws_freq = wb.create_sheet("FRÉQUENCE", 1)

        ws_freq['A1'] = "FRÉQUENCE D'APPARITION DES NUMÉROS"
        ws_freq['A1'].font = title_font
        ws_freq.merge_cells('A1:E1')

        headers = ["NUMÉRO", "NB JOURS", "% JOURS", "JOURS PRÉSENTS", "CATÉGORIES"]
        for col, header in enumerate(headers, 1):
            cell = ws_freq.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Récupérer fréquences
        rows = await conn.fetch("""
            SELECT 
                game_number,
                COUNT(DISTINCT jour_id) as nb_jours,
                array_agg(DISTINCT jour_id ORDER BY jour_id) as jours_list,
                array_agg(DISTINCT category) as categories
            FROM games
            GROUP BY game_number
            ORDER BY nb_jours DESC, game_number ASC
        """)

        total_jours = len(jours_list)

        for idx, row in enumerate(rows, 4):
            ws_freq.cell(row=idx, column=1, value=row['game_number'])
            ws_freq.cell(row=idx, column=2, value=row['nb_jours'])
            ws_freq.cell(row=idx, column=3, value=f"{(row['nb_jours']/total_jours)*100:.1f}%")
            ws_freq.cell(row=idx, column=4, value=", ".join(row['jours_list']))
            ws_freq.cell(row=idx, column=5, value=", ".join(row['categories']))

            # Colorer si présent dans tous les jours
            if row['nb_jours'] == total_jours:
                for col in range(1, 6):
                    ws_freq.cell(row=idx, column=col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_freq.column_dimensions[col].width = 18

        # === FEUILLE 3: NUMÉROS COMMUNS ===
        ws_common = wb.create_sheet("NUMÉROS COMMUNS", 2)

        ws_common['A1'] = f"NUMÉROS PRÉSENTS DANS TOUTES LES JOURNÉES ({total_jours} jours)"
        ws_common['A1'].font = title_font
        ws_common.merge_cells('A1:D1')

        headers = ["NUMÉRO", "CATÉGORIE", "CONSTANT", "DÉTAIL"]
        for col, header in enumerate(headers, 1):
            cell = ws_common.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Numéros dans tous les jours avec même catégorie
        common = await conn.fetch("""
            SELECT game_number, category, COUNT(*) as freq
            FROM games
            GROUP BY game_number, category
            HAVING COUNT(DISTINCT jour_id) = $1
            ORDER BY game_number
        """, total_jours)

        for idx, row in enumerate(common, 4):
            ws_common.cell(row=idx, column=1, value=row['game_number'])
            ws_common.cell(row=idx, column=2, value=row['category'])
            ws_common.cell(row=idx, column=3, value="✅ OUI")
            ws_common.cell(row=idx, column=4, value=f"Présent dans {row['freq']} fois")

        ws_common.column_dimensions['A'].width = 12
        ws_common.column_dimensions['B'].width = 12
        ws_common.column_dimensions['C'].width = 12
        ws_common.column_dimensions['D'].width = 20

        # === FEUILLE 4: MATRICE JOURNÉES ===
        ws_matrix = wb.create_sheet("MATRICE", 3)

        ws_matrix['A1'] = "MATRICE DE PRÉSENCE DES NUMÉROS PAR JOURNÉE"
        ws_matrix['A1'].font = title_font
        ws_matrix.merge_cells('A1:H1')

        # En-têtes: Numéro + une colonne par jour
        ws_matrix['A3'] = "NUMÉRO"
        ws_matrix['A3'].font = header_font
        ws_matrix['A3'].fill = header_fill

        for idx, jour in enumerate(jours_list, 2):
            col_letter = get_column_letter(idx)
            ws_matrix[f'{col_letter}3'] = jour[-5:]  # Derniers 5 caractères (MM-DD)
            ws_matrix[f'{col_letter}3'].font = header_font
            ws_matrix[f'{col_letter}3'].fill = header_fill
            ws_matrix[f'{col_letter}3'].alignment = Alignment(horizontal="center", text_rotation=45)

        # Récupérer tous les numéros uniques
        all_numbers = await conn.fetch("""
            SELECT DISTINCT game_number FROM games ORDER BY game_number
        """)

        # Pour chaque numéro, voir dans quels jours il est présent
        for row_idx, num_row in enumerate(all_numbers, 4):
            num = num_row['game_number']
            ws_matrix.cell(row=row_idx, column=1, value=num)

            for col_idx, jour in enumerate(jours_list, 2):
                exists = await conn.fetchval("""
                    SELECT 1 FROM games WHERE game_number = $1 AND jour_id = $2 LIMIT 1
                """, num, jour)

                col_letter = get_column_letter(col_idx)
                if exists:
                    ws_matrix[f'{col_letter}{row_idx}'] = "✓"
                    ws_matrix[f'{col_letter}{row_idx}'].alignment = Alignment(horizontal="center")
                    ws_matrix[f'{col_letter}{row_idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    ws_matrix[f'{col_letter}{row_idx}'] = ""

        ws_matrix.column_dimensions['A'].width = 10
        for idx in range(2, len(jours_list) + 2):
            ws_matrix.column_dimensions[get_column_letter(idx)].width = 8

        # === FEUILLE 5: STATISTIQUES PAR CATÉGORIE ===
        ws_cat = wb.create_sheet("STATS CATÉGORIES", 4)

        ws_cat['A1'] = "RÉPARTITION PAR CATÉGORIE SUR TOUTES LES JOURNÉES"
        ws_cat['A1'].font = title_font
        ws_cat.merge_cells('A1:E1')

        headers = ["CATÉGORIE", "TOTAL", "MOYENNE/JOUR", "MIN", "MAX"]
        for col, header in enumerate(headers, 1):
            cell = ws_cat.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        categories = ['✅0️⃣', '✅1️⃣', '✅2️⃣', '✅3️⃣', '❌']

        for idx, cat in enumerate(categories, 4):
            stats = await conn.fetchrow("""
                SELECT 
                    COUNT(*) as total,
                    AVG(COUNT) as moyenne,
                    MIN(COUNT) as min_val,
                    MAX(COUNT) as max_val
                FROM (
                    SELECT jour_id, COUNT(*) as COUNT
                    FROM games 
                    WHERE category = $1
                    GROUP BY jour_id
                ) sub
            """, cat)

            ws_cat.cell(row=idx, column=1, value=cat)
            ws_cat.cell(row=idx, column=2, value=stats['total'])
            ws_cat.cell(row=idx, column=3, value=f"{stats['moyenne']:.1f}")
            ws_cat.cell(row=idx, column=4, value=stats['min_val'])
            ws_cat.cell(row=idx, column=5, value=stats['max_val'])

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_cat.column_dimensions[col].width = 15

        filename = f"comparaison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        return filename, None

@client.on(events.NewMessage(pattern='/compare_all'))
async def cmd_compare_all(event):
    """Génère un fichier Excel UNIQUEMENT avec les analyses comparatives"""
    if event.is_group or event.is_channel:
        return

    await event.respond("📊 Génération du fichier de comparaison...")

    try:
        filename, error = await create_comparison_only_excel()

        if error:
            await event.respond(f"❌ {error}")
            return

        await client.send_file(
            event.chat_id,
            filename,
            caption=f"""📊 **FICHIER DE COMPARAISON GLOBALE**

📁 **Contenu:**
• **RÉCAP GLOBAL** - Vue d'ensemble des journées
• **FRÉQUENCE** - Combien de fois chaque numéro apparaît
• **NUMÉROS COMMUNS** - Présents dans TOUTES les journées
• **MATRICE** - Tableau de présence jour par jour
• **STATS CATÉGORIES** - Moyennes et extrêmes

💡 Les cellules vertes = numéros présents dans toutes les journées"""
        )

        os.remove(filename)

    except Exception as e:
        logger.error(f"Erreur comparaison: {e}")
        await event.respond(f"❌ Erreur: {e}")
@client.on(events.NewMessage(pattern='/hot_numbers'))
async def cmd_hot_numbers(event):
    if event.is_group or event.is_channel:
        return

    try:
        data = await get_global_comparison_data()

        if len(data['jours']) < 2:
            await event.respond("❌ Pas assez de journées pour analyser")
            return

        # Numéros présents dans au moins la moitié des jours
        min_days = len(data['jours']) // 2
        hot_numbers = [n for n in data['number_freq'] if n['nb_jours'] >= min_days]

        msg = f"""🔥 **NUMÉROS CHAUDS** (présents dans {min_days}+ jours)

**Total journées:** {len(data['jours'])}
**Numéros fréquents:** {len(hot_numbers)}

**Top numéros:**
"""
        for row in hot_numbers[:30]:  # Limiter à 30
            msg += f"\n• **{row['game_number']}**: présent dans {row['nb_jours']} jours"

        await event.respond(msg)

    except Exception as e:
        await event.respond(f"❌ Erreur: {e}")

# Serveur Web
async def health_check(request):
    return web.Response(text="OK", status=200)

async def start_web_server():
    app = web.Application()
    app.router.add_get('/health', health_check)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, '0.0.0.0', PORT)
    await site.start()
    logger.info(f"Serveur web démarré sur port {PORT}")

# Démarrage
client.add_event_handler(handle_edited_message, events.MessageEdited())

async def main():
    # Connexion PostgreSQL
    await db.connect()

    # Démarrer serveur web
    await start_web_server()

    # Démarrer Telegram
    await client.start(bot_token=BOT_TOKEN)
    logger.info("Bot Telegram connecté")

    # Démarrer planificateur
    asyncio.create_task(bilan_scheduler())
    logger.info(f"Bilans automatiques: {bilan_interval} min")

    logger.info("🚀 Bot opérationnel!")
    await client.run_until_disconnected()

if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Bot arrêté")
    except Exception as e:
        logger.error(f"Erreur: {e}")
